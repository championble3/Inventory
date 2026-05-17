from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker
from backend.models.bm32 import DrawingRecord, engine

Session = sessionmaker(bind=engine)
session = Session()

df = pd.read_excel(r"F:\REGENERACJA - DOKUMENTACJA\Rejestr rysunków.xlsm", sheet_name='Dno formy')
ws = load_workbook(r"F:\REGENERACJA - DOKUMENTACJA\Rejestr rysunków.xlsm")['Dno formy']

# xls = pd.ExcelFile(r"F:\REGENERACJA - DOKUMENTACJA\Rejestr rysunków.xlsm")
# print(xls.sheet_names)
def ingestion_func(df, ws):
    ws_rows = list(ws.iter_rows(min_row=2))
    
    for (_, row), ws_row in zip(df.iterrows(), ws_rows):
        pdf_cell   = ws_row[4]
        pliki_cell = ws_row[5]

        # Wyczyść datę — NaN zamień na None
        raw_date = row['Data']
        clean_date = None if pd.isna(raw_date) else raw_date

        # Wyczyść materiał — liczby zamień na string
        raw_material = row['Materiał']
        clean_material = None if pd.isna(raw_material) else str(raw_material)

        new_record = DrawingRecord(
            nr_rys    = row['Nr rys.'],
            full_name = row['Pełna nazwa'],
            material  = clean_material,
            date      = clean_date,
            pdf_url   = pdf_cell.hyperlink.target   if pdf_cell.hyperlink   else None,
            pliki_url = pliki_cell.hyperlink.target if pliki_cell.hyperlink else None,
        )
       #session.add(new_record)

    # session.commit()
    # print(f"Dodano {len(df)} rekordów")
    print(new_record)

ingestion_func(df,ws)

# print("=== WSZYSTKIE REKORDY ===")
# records = session.query(BM32).all()
# for record in records:
#     print(record)

# record = session.query(BM32).count()

# print(record)