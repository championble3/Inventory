from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session, sessionmaker
from typing import List
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import io
import os
import platform
import subprocess
from pathlib import Path
from urllib.parse import unquote
from dotenv import load_dotenv

from models.bm32 import BM32, engine
from pydantic import BaseModel


# ==================== Konfiguracja ścieżek ====================

# Wczytaj zmienne środowiskowe z pliku .env
load_dotenv()

BASE_DOCS_PATH = os.getenv("BASE_DOCS_PATH", r"F:\REGENERACJA - DOKUMENTACJA")

SessionLocal = sessionmaker(bind=engine)


def get_db():
    """Dependency do pobierania sesji bazy danych"""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ==================== Pydantic Schematy ====================

class BM32Base(BaseModel):
    """Bazowy schemat BM32"""
    nr_rys: str
    full_name: str | None = None
    material: str | None = None
    pdf_url: str | None = None
    pliki_url: str | None = None


class BM32Create(BM32Base):
    """Schemat do tworzenia nowego rekordu"""
    pass


class BM32Update(BaseModel):
    """Schemat do aktualizacji rekordu"""
    full_name: str | None = None
    material: str | None = None
    pdf_url: str | None = None
    pliki_url: str | None = None


class BM32Response(BM32Base):
    """Schemat odpowiedzi"""
    date: datetime

    class Config:
        from_attributes = True


class IngestionResult(BaseModel):
    """Schemat wyniku ingestion"""
    success: bool
    records_added: int
    message: str


# ==================== Router ====================

router = APIRouter(prefix="/api/bm32", tags=["BM32"])


# ==================== GET Endpoints ====================
#add limit = 100
@router.get("/", response_model=List[BM32Response])
def get_all_bm32(skip: int = 0, db: Session = Depends(get_db)):
    """
    Pobierz wszystkie rekordy BM32
    
    - **skip**: liczba rekordów do pominięcia
    - **limit**: maksymalna liczba rekordów do zwrócenia
    """
    bm32_records = db.query(BM32).offset(skip).all()
    return bm32_records


@router.get("/{nr_rys}", response_model=BM32Response)
def get_bm32_by_id(nr_rys: str, db: Session = Depends(get_db)):
    """
    Pobierz rekord BM32 po numerze rysunku
    
    - **nr_rys**: numer rysunku
    """
    bm32_record = db.query(BM32).filter(BM32.nr_rys == nr_rys).first()
    
    if not bm32_record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rekord z nr_rys={nr_rys} nie znaleziony"
        )
    
    return bm32_record


# ==================== POST Endpoints ====================

@router.post("/", response_model=BM32Response, status_code=status.HTTP_201_CREATED)
def create_bm32(bm32: BM32Create, db: Session = Depends(get_db)):
    """
    Utwórz nowy rekord BM32
    
    - **nr_rys**: numer rysunku (wymagane)
    - **full_name**: pełna nazwa
    - **material**: materiał
    - **pdf_url**: URL do PDF
    - **pliki_url**: URL do plików
    """
    # Sprawdzenie czy rekord już istnieje
    existing_record = db.query(BM32).filter(BM32.nr_rys == bm32.nr_rys).first()
    if existing_record:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Rekord z nr_rys={bm32.nr_rys} już istnieje"
        )
    
    # Tworzenie nowego rekordu
    new_bm32 = BM32(
        nr_rys=bm32.nr_rys,
        full_name=bm32.full_name,
        material=bm32.material,
        pdf_url=bm32.pdf_url,
        pliki_url=bm32.pliki_url,
        date=datetime.now()
    )
    
    db.add(new_bm32)
    db.commit()
    db.refresh(new_bm32)
    
    return new_bm32


# ==================== PUT/PATCH Endpoints ====================

@router.put("/{nr_rys}", response_model=BM32Response)
def update_bm32(nr_rys: str, bm32_update: BM32Update, db: Session = Depends(get_db)):
    """
    Aktualizuj rekord BM32
    
    - **nr_rys**: numer rysunku
    """
    bm32_record = db.query(BM32).filter(BM32.nr_rys == nr_rys).first()
    
    if not bm32_record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rekord z nr_rys={nr_rys} nie znaleziony"
        )
    
    # Aktualizacja pól
    update_data = bm32_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(bm32_record, field, value)
    
    db.commit()
    db.refresh(bm32_record)
    
    return bm32_record


# ==================== DELETE Endpoints ====================

@router.delete("/{nr_rys}", status_code=status.HTTP_204_NO_CONTENT)
def delete_bm32(nr_rys: str, db: Session = Depends(get_db)):
    """
    Usuń rekord BM32
    
    - **nr_rys**: numer rysunku
    """
    bm32_record = db.query(BM32).filter(BM32.nr_rys == nr_rys).first()
    
    if not bm32_record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rekord z nr_rys={nr_rys} nie znaleziony"
        )
    
    db.delete(bm32_record)
    db.commit()
    
    return None


# ==================== Search/Filter Endpoints ====================

@router.get("/search/material/{material}", response_model=List[BM32Response])
def search_by_material(material: str, db: Session = Depends(get_db)):
    """
    Wyszukaj rekordy po materiale
    
    - **material**: materiał do wyszukania
    """
    records = db.query(BM32).filter(BM32.material.contains(material)).all()
    return records


@router.get("/search/name/{full_name}", response_model=List[BM32Response])
def search_by_name(full_name: str, db: Session = Depends(get_db)):
    """
    Wyszukaj rekordy po nazwie
    
    - **full_name**: nazwa do wyszukania
    """
    records = db.query(BM32).filter(BM32.full_name.contains(full_name)).all()
    return records


# ==================== Ingestion Endpoint ====================

@router.post("/ingestion/excel", response_model=IngestionResult)
def ingest_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Wczytaj dane z pliku Excel
    
    Oczekuje pliku .xlsx z arkuszem 'BM32' zawierającym kolumny:
    - Nr rys.
    - Pełna nazwa
    - Materiał
    - Data
    - pdf_url (hiperlącze)
    - pliki_url (hiperlącze)
    """
    try:
        # Odczyt pliku Excel
        contents = file.file.read()
        excel_file = io.BytesIO(contents)
        
        # Załaduj dane
        df = pd.read_excel(excel_file, sheet_name='BM32')
        excel_file.seek(0)
        ws = load_workbook(excel_file)['BM32']
        
        ws_rows = list(ws.iter_rows(min_row=2))
        records_added = 0
        
        for (_, row), ws_row in zip(df.iterrows(), ws_rows):
            # Pobierz hiperłącza
            pdf_cell = ws_row[4]
            pliki_cell = ws_row[5]
            
            # Wyczyść datę
            raw_date = row.get('Data')
            clean_date = None if pd.isna(raw_date) else raw_date
            
            # Wyczyść materiał
            raw_material = row.get('Materiał')
            clean_material = None if pd.isna(raw_material) else str(raw_material)
            
            # Sprawdź czy rekord już istnieje
            nr_rys = row.get('Nr rys.')
            existing = db.query(BM32).filter(BM32.nr_rys == nr_rys).first()
            
            if not existing:
                new_record = BM32(
                    nr_rys=nr_rys,
                    full_name=row.get('Pełna nazwa'),
                    material=clean_material,
                    date=clean_date,
                    pdf_url=pdf_cell.hyperlink.target if pdf_cell.hyperlink else None,
                    pliki_url=pliki_cell.hyperlink.target if pliki_cell.hyperlink else None,
                )
                db.add(new_record)
                records_added += 1
        
        db.commit()
        
        return IngestionResult(
            success=True,
            records_added=records_added,
            message=f"Pomyślnie wczytano {records_added} nowych rekordów z pliku {file.filename}"
        )
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Błąd podczas ingestion: {str(e)}"
        )


# ==================== File Management Endpoints ====================

@router.post("/{nr_rys}/upload-pdf")
async def upload_pdf(nr_rys: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Prześlij plik PDF dla rekordu
    
    - **nr_rys**: numer rysunku
    - **file**: plik PDF
    """
    bm32_record = db.query(BM32).filter(BM32.nr_rys == nr_rys).first()
    
    if not bm32_record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rekord z nr_rys={nr_rys} nie znaleziony"
        )
    
    try:
        # Utwórz folder jeśli nie istnieje
        pdfs_dir = Path(__file__).parent.parent / "pdfs" / nr_rys
        pdfs_dir.mkdir(parents=True, exist_ok=True)
        
        # Zapisz plik
        file_path = pdfs_dir / file.filename
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
        
        # Aktualizuj ścieżkę w bazie
        bm32_record.pdf_url = str(file_path)
        db.commit()
        
        return {
            "success": True,
            "message": f"PDF {file.filename} załadowany pomyślnie",
            "pdf_path": str(file_path)
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Błąd przy przesyłaniu PDF: {str(e)}"
        )


@router.get("/{nr_rys}/open-pdf")
def open_pdf(nr_rys: str, db: Session = Depends(get_db)):
    """
    Otwórz plik PDF dla rekordu
    
    - **nr_rys**: numer rysunku
    """
    bm32_record = db.query(BM32).filter(BM32.nr_rys == nr_rys).first()
    
    if not bm32_record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rekord z nr_rys={nr_rys} nie znaleziony"
        )
    
    if not bm32_record.pdf_url:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Brak przypisanego pliku PDF dla nr_rys={nr_rys}"
        )
    
    # Zdekoduj URL-encoded ścieżkę
    pdf_path = unquote(bm32_record.pdf_url)
    if not os.path.isabs(pdf_path):
        pdf_path = os.path.join(BASE_DOCS_PATH, pdf_path)
    
    try:
        if not os.path.exists(pdf_path):
            raise Exception(f"Plik nie istnieje na dysku: {pdf_path}")
        
        # Otwórz plik w domyślnej aplikacji
        if platform.system() == 'Windows':
            os.startfile(pdf_path)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', pdf_path])
        else:  # Linux
            subprocess.run(['xdg-open', pdf_path])
        
        return {"success": True, "message": f"Otwieranie pliku: {pdf_path}"}
    
    except Exception as e:
        print(f"ERROR: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Błąd przy otwieraniu PDF: {str(e)}"
        )


@router.get("/{nr_rys}/open-files")
def open_files(nr_rys: str, db: Session = Depends(get_db)):
    """
    Otwórz folder z plikami dla rekordu
    
    - **nr_rys**: numer rysunku
    """
    bm32_record = db.query(BM32).filter(BM32.nr_rys == nr_rys).first()
    
    if not bm32_record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Rekord z nr_rys={nr_rys} nie znaleziony"
        )
    
    if not bm32_record.pliki_url:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Brak przypisanego folderu z plikami dla nr_rys={nr_rys}"
        )
    
    # Zdekoduj URL-encoded ścieżkę
    files_path = unquote(bm32_record.pliki_url)
    if not os.path.isabs(files_path):
        files_path = os.path.join(BASE_DOCS_PATH, files_path)
    
    try:
        if not os.path.exists(files_path):
            raise Exception(f"Folder nie istnieje na dysku: {files_path}")
        
        # Otwórz folder w eksploratorze
        if platform.system() == 'Windows':
            os.startfile(files_path)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', files_path])
        else:  # Linux
            subprocess.run(['xdg-open', files_path])
        
        return {"success": True, "message": f"Otwieranie folderu: {files_path}"}
    
    except Exception as e:
        print(f"ERROR: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Błąd przy otwieraniu folderu: {str(e)}"
        )



