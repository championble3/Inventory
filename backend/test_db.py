import sys
from pathlib import Path
import pandas as pd
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook


# Dodaj backend do ścieżki
sys.path.insert(0, str(Path(__file__).parent.parent))

from backend.models.bm32 import BM32, engine

Session = sessionmaker(bind=engine)
session = Session()

# === ODCZYT - Po nr_rys ===
# print("\n=== SZUKAJ PO nr_rys ===")
# record = session.query(BM32).filter(BM32.nr_rys == 1).first()
# if record:
#     print(record)
# else:
#     print("Nie znaleziono")

# === DODAJ NOWY REKORD ===
# print("\n=== DODAJ NOWY ===")
# new_record = BM32(
#     nr_rys=100,
#     full_name="Test Produkt",
#     material="Stal",
#     pdf_url="http://example.com/test.pdf"
# )
# session.add(new_record)
# session.commit()
# print(f"Dodano: {new_record}")

# # === EDYTUJ ===
# print("\n=== EDYTUJ ===")
# record = session.query(BM32).filter(BM32.nr_rys == 100).first()
# if record:
#     record.material = "Aluminium"
#     session.commit()
#     print(f"Zmieniono: {record}")

# # === USUŃ ===
# print("\n=== USUŃ ===")
# record = session.query(BM32).filter(BM32.nr_rys == 101).first()
# if record:
#     session.delete(record)
#     session.commit()
#     print(f"Usunięto {record}")

# session.close()

# DB FILLING
df = pd.read_excel(r"F:\REGENERACJA - DOKUMENTACJA\Rejestr rysunków.xlsm", sheet_name='BM32')
ws = load_workbook(r"F:\REGENERACJA - DOKUMENTACJA\Rejestr rysunków.xlsm")['BM32']

def ingestion_func(df, ws):
    ws_rows = list(ws.iter_rows(min_row=2))
    
    for (_, row), ws_row in zip(df.iterrows(), ws_rows):
        pdf_cell   = ws_row[4]
        pliki_cell = ws_row[5]

        # Wyczyść datę — NaN zamień na None
        raw_date = row['Data']
        clean_date = None if pd.isna(raw_date) else raw_date

        # Wyczyść materiał — liczby zamień na string
        raw_material = row['Materiał']
        clean_material = None if pd.isna(raw_material) else str(raw_material)

        new_record = BM32(
            nr_rys    = row['Nr rys.'],
            full_name = row['Pełna nazwa'],
            material  = clean_material,
            date      = clean_date,
            pdf_url   = pdf_cell.hyperlink.target   if pdf_cell.hyperlink   else None,
            pliki_url = pliki_cell.hyperlink.target if pliki_cell.hyperlink else None,
        )
        session.add(new_record)

    session.commit()
    print(f"Dodano {len(df)} rekordów")

#ingestion_func(df,ws)

print("=== WSZYSTKIE REKORDY ===")
records = session.query(BM32).all()
for record in records:
    print(record)
